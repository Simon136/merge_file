# -*- coding: utf-8 -*-

import zipfile, gzip, tarfile, shutil, xlrd
import pandas as pd
import numpy
import openpyxl
import os
from PyQt5.QtCore import QThread, pyqtSignal
from datetime import datetime


class New_Thread(QThread):
    finishSignal = pyqtSignal(dict)

    def __init__(self, merge_path_file, save_file_path, row_content, header_type, template_path, first_use, unzip_folder, final_row, parent=None):
        """
        初始化传入参数
        :param merge_path_file: 传入需合并文件夹或压缩包
        :param save_file_path: 文件或压缩包保存路径
        :param row_content: 选择表头指针
        :param header_type: 表头选择类型
        :param template_path: 模板路径
        :param parent: None
        """
        super(New_Thread, self).__init__(parent)
        self.merge_path_file = merge_path_file
        self.save_file_path = save_file_path
        self.row_content = row_content
        self.header_type = header_type
        self.template_path = template_path
        self.final_row = final_row
        if os.path.isdir(self.merge_path_file):
            print("合并类型为文件夹输入")
        else:
            if first_use:
                print("合并类型为压缩包输入")
                fuglog = decompression_process(self.merge_path_file, self.save_file_path)
                save_file_now = fuglog.decompressing_files()
                self.merge_path_file = save_file_now
            else:
                self.merge_path_file = unzip_folder

    def run(self):
        try:
            feedback_start = self.template_check()
            self.finishSignal.emit(feedback_start)
        except Exception as e:
            feedback_start = {'状态': False, '内容': str(e)}
            self.finishSignal.emit(feedback_start)

    def template_check(self):
        feedback_start = {}
        if self.template_path == "":
            excel_name = os.listdir(self.merge_path_file)[0]
            self.template_path = self.merge_path_file+'\\'+excel_name
        if self.header_type == "按标题名称合并":
            file_extension = os.path.splitext(self.template_path)[1]
            if file_extension == '.xlsx':
                workbook = openpyxl.load_workbook(self.template_path)
                worksheet = workbook.active
                min_row = worksheet.min_row
                max_row = worksheet.max_row
                min_col = worksheet.min_column
                max_col = worksheet.max_column
            elif file_extension == '.xls':
                workbook = xlrd.open_workbook(self.template_path)
                worksheet = workbook.sheet_by_index(0)
                max_row = worksheet.nrows
                max_col = worksheet.ncols
                min_row = 0
                min_col = 0
            elif file_extension == '.csv':
                feedback_start['状态'] = False
                feedback_start['内容'] = "csv格式暂不支持按标题名称进行合并，请找到其对应的标题所在行，使用标题所在行合并"
                return feedback_start
            else:
                feedback_start['状态'] = False
                feedback_start['内容'] = "无法处理对应格式"
                return feedback_start

            row = 0
            judge = False
            print(self.row_content)
            for m in range(min_row, max_row):
                for j in range(min_col, max_col):
                    cell = worksheet.cell(m, j).value
                    if cell == self.row_content:
                        judge = True
                        break
                if judge:
                    break
                else:
                    row = row+1
            if not judge:
                feedback_start['状态'] = False
                feedback_start['内容'] = "无法找到对应的指定内容，请核查是否输入有误"
            else:
                data_header = pd.read_excel(self.template_path, header=row)
                header_content = ' | '.join(list(data_header.columns))
                feedback_start['状态'] = True
                feedback_start['内容'] = header_content
                feedback_start['表头内容'] = list(data_header.columns)
                feedback_start['所在行'] = row
                feedback_start['文件所在文件夹'] = self.merge_path_file
        else:
            row = int(self.row_content)
            row = row-1
            file_extension = os.path.splitext(self.template_path)[1]
            if file_extension == '.xlsx' or file_extension == '.xls':
                data_header = pd.read_excel(self.template_path, header=row)
                header_content = ' | '.join(list(data_header.columns))
                feedback_start['状态'] = True
                feedback_start['内容'] = header_content
                feedback_start['表头内容'] = list(data_header.columns)
                feedback_start['所在行'] = row
                feedback_start['文件所在文件夹'] = self.merge_path_file
            elif file_extension == '.csv':
                data_header = pd.read_csv(self.template_path, header=row, encoding='gbk')
                header_content = ' | '.join(list(data_header.columns))
                feedback_start['状态'] = True
                feedback_start['内容'] = header_content
                feedback_start['表头内容'] = list(data_header.columns)
                feedback_start['所在行'] = row
                feedback_start['文件所在文件夹'] = self.merge_path_file
            else:
                feedback_start['状态'] = False
                feedback_start['内容'] = "无法处理对应格式"
        return feedback_start


class decompression_process():
    '''
    文件解压
    '''
    def __init__(self, file_name, save_path):
        """
        文件解压
        :param file_name: 文件路径
        :param save_path: 保存路径
        """
        self.file_name = file_name
        self.save_path = save_path
        self.file_name = self.file_name.replace('/', '\\')
        print(self.file_name)
        print(self.save_path)

    def decompressing_files(self):
        """
        解压文件汇总函数
        :return: 解压文件夹
        """
        name_suffix = os.path.splitext(self.file_name)[-1]
        if name_suffix == ".zip":
            save_file_now = self.un_zip(self.file_name, self.save_path)
        elif name_suffix == ".gz":
            save_file_now = self.un_gz(self.file_name, self.save_path)
        elif name_suffix == ".tar":
            save_file_now = self.un_tar(self.file_name, self.save_path)
        else:
            save_file_now = "无对应解压类型处理"
        return save_file_now

    def un_gz(self, file_name, save_path):
        """
        解压gz文件
        :return: 解压文件路径
        """
        save_file = file_name.replace(".gz", "").split("\\")[-1]
        save_file = save_path + '\\' + save_file
        print(save_file)
        with gzip.open(self.file_name, 'rb') as f_in:
            with open(save_file, 'wb') as f_out:
                shutil.copyfileobj(f_in, f_out)
        return save_file

    def un_tar(self, file_name, save_path):
        """
        解压tar文件
        :param file_name: 文件路径
        :param save_path: 解压路径
        :return: 解压文件路径
        """
        tar = tarfile.open(file_name)
        names = tar.getnames()
        save_name = file_name.replace(".tar", "").split("\\")[-1]
        save_file = save_path + '\\' + save_name
        save_file_now = self.create_unique_folder(save_file)
        for name in names:
            tar.extract(name, save_file_now + "/")
        tar.close()
        return save_file_now

    def un_zip(self, file_name, save_path):
        """
        解压zip文件
        :return: 解压文件路径
        """
        save_name = file_name.replace(".zip", "").split("\\")[-1]
        save_file = save_path + '\\' + save_name
        save_file_now = self.create_unique_folder(save_file)
        with zipfile.ZipFile(self.file_name) as f:
            for name in f.namelist()[::-1]:
                f.extract(name, save_file_now + "/")
                os.rename(save_file_now + '\\' + name, save_file_now + '\\' + name.encode('cp437').decode('gbk'))
        return save_file_now

    def create_unique_folder(self, file_path):
        """
        判断文件夹是否存在，若存在则新建文件夹
        :param file_path: 文件夹
        :return:文件夹路径
        """
        i = 1
        file_path_now = file_path
        while os.path.isdir(file_path_now):
            file_path_now = file_path+"("+str(i)+")"
            i = i + 1
        os.mkdir(file_path_now)
        return file_path_now


class Summary_thread(QThread):
    finishSignal = pyqtSignal(dict)

    def __init__(self, header_template, row_final, save_path, merge_path_path, parent=None):
        """
        汇总文件函数调用
        :param header_template: 汇总函数表头
        :param row_final: 汇总函数所在行
        :param save_path: 汇总函数保存路径
        :param merge_path_path: 汇总函数目标汇总文件夹
        :param parent: None
        """
        super(Summary_thread, self).__init__(parent)
        self.header_template = header_template
        self.row_final = row_final
        self.save_path = save_path
        self.merge_path_path = merge_path_path.replace('/', '\\')
        self.return_dic = {}

    def run(self):
        try:
            file_get_data = file_merge_class(self.header_template, self.row_final, self.save_path, self.merge_path_path)
            data_dic = file_get_data.file_merge_excel()
            self.return_dic['状态'] = True
            self.return_dic['内容'] = data_dic
            self.finishSignal.emit(self.return_dic)
        except Exception as e:
            error = str(e)
            self.return_dic['状态'] = False
            self.return_dic['内容'] = error
            self.finishSignal.emit(self.return_dic)


class file_merge_class():
    def __init__(self, header_template, row_final, save_path, merge_path_path):
        """
        文件合并汇总
        :param header_template: 表头数组
        :param row_final: 所取行位置
        :param save_path: 保存位置
        :param merge_path_path: 读取文件夹位置
        """
        self.header_template = header_template
        self.row_final = row_final
        self.save_path = save_path
        self.merge_path_path = merge_path_path
        # 设计异常无法汇总文件
        self.error_save_file = save_path+'\\'+merge_path_path.split('\\')[-1]+'_异常无法汇总文件'
        # 创建文件夹，避免创建多个重复文件夹
        self.error_save_file = self.create_unique_folder(self.error_save_file)
        self.unable_merge_dic = {}
        self.sum_dataframe = pd.DataFrame()

    def file_merge_excel(self):
        feedback_start = {}
        for file_name in os.listdir(self.merge_path_path):
            file_extension = os.path.splitext(file_name)[1]
            excel_name = self.merge_path_path+'\\'+file_name
            if file_extension == '.xlsx' or file_extension == '.xls':
                self.read_excel_function(excel_name, file_extension)
            elif file_extension == '.csv':
                data = pd.read_csv(excel_name, header=self.row_final, encoding='gbk')
                if len(data) != 0:
                    data_header = data.columns
                    if list(data_header) != self.header_template:
                        self.unable_merge_dic[os.path.splitext(file_name)[0]] = excel_name
                    else:
                        self.sum_dataframe = pd.concat([data, self.sum_dataframe])
            else:
                self.unable_merge_dic[os.path.splitext(file_name)[0]] = excel_name
        for key_name in self.unable_merge_dic:
            shutil.move(self.unable_merge_dic[key_name], self.error_save_file)
        unable_table = pd.DataFrame(list(self.unable_merge_dic.items()), columns=['sheet页表', '文件路径'])
        unable_table_save = self.error_save_file+'\\'+'无法合并表格记录表.xlsx'
        unable_table.to_excel(unable_table_save, index=False)
        time_now = datetime.now()
        time_now = time_now.strftime('%Y-%m-%d_%H%M%S')
        save_name = self.merge_path_path.split('\\')[-1]
        save_path_sumdataframe = self.save_path+'\\'+time_now+'_'+save_name+'.xlsx'
        self.sum_dataframe.to_excel(save_path_sumdataframe, index=False)
        feedback_start['汇总表'] = save_path_sumdataframe
        feedback_start['无法汇总记录表'] = unable_table_save
        feedback_start['无法汇总文件迁移'] = self.error_save_file
        return feedback_start

    def read_excel_function(self, excel_name, try_excel):
        xlsx_sheet = pd.ExcelFile(excel_name)
        sheet_names = xlsx_sheet.sheet_names
        sum_dataframe = pd.DataFrame()
        for sheet_name in sheet_names:
            if try_excel == '.xlsx':
                data = pd.read_excel(excel_name, header=self.row_final, sheet_name=sheet_name, engine='openpyxl')
            else:
                data = pd.read_excel(excel_name, header=self.row_final, sheet_name=sheet_name, engine='xlrd')
            if len(data) != 0:
                data_header = data.columns
                if list(data_header) != self.header_template:
                    self.unable_merge_dic[sheet_name] = excel_name
                    return False
                else:
                    sum_dataframe = pd.concat([data, sum_dataframe])
        self.sum_dataframe = pd.concat([sum_dataframe, self.sum_dataframe])
        return True

    def create_unique_folder(self, file_path):
        """
        判断文件夹是否存在，若存在则新建文件夹
        :param file_path: 文件夹
        :return:文件夹路径
        """
        i = 1
        file_path_now = file_path
        while os.path.isdir(file_path_now):
            file_path_now = file_path+"("+str(i)+")"
            i = i + 1
        os.mkdir(file_path_now)
        return file_path_now


if __name__ == "__main__":
    merge_path_file = r"D:\1.项目文件\yindao\即刻文化-自助分析\1.RPA自动化处理\1.下载全数据\2023-11-17"
    save_file_path = r"D:\1.项目文件\yindao\即刻文化-自助分析\1.RPA自动化处理\1.下载全数据"
    row_content = '日期'
    header_type = "按标题名称合并"
    template_path = ""
    # un_tar(file_name, save_path)
    # un_zip(file_name, save_path)
    # fuglog = decompression_process(file_name, save_path)
    # save_file_now = fuglog.decompressing_files()
    # print(save_file_now)
    # fuglog = New_Thread(merge_path_file, save_file_path, row_content, header_type, template_path)
    # fuglog.template_check()
    header_template = ['统计日期', '访客数', '浏览量', '商品访客数', '商品浏览量', '平均停留时长', '跳失率', '商品收藏买家数',
       '商品收藏次数', '加购人数', '支付金额', '支付买家数', '支付子订单数', '支付件数', '下单金额', '下单买家数',
       '下单件数', '人均浏览量', '下单转化率', '支付转化率', '客单价', 'UV价值', '老访客数', '新访客数',
       '加购件数', '支付老买家数', '老买家支付金额', '直通车消耗', '钻石展位消耗', '淘宝客佣金', '成功退款金额',
       '评价数', '有图评价数', '正面评价数', '负面评价数', '老买家正面评价数', '老买家负面评价数', '支付父订单数',
       '揽收包裹数', '发货包裹数', '派送包裹数', '签收成功包裹数', '平均支付_签收时长(秒)', '描述相符评分',
       '物流服务评分', '服务态度评分', '下单-支付转化率', '支付商品数', '店铺收藏买家数']
    row_final = 7
    save_path = r'D:\1.项目文件\yindao\即刻文化-自助分析\1.RPA自动化处理\1.下载全数据'
    merge_path_path = r'D:\1.项目文件\yindao\即刻文化-自助分析\1.RPA自动化处理\1.下载全数据\2023-11-22'
    file_get_data = file_merge_class(header_template, row_final, save_path, merge_path_path)
    data_save = file_get_data.file_merge_excel()
    print(data_save)
